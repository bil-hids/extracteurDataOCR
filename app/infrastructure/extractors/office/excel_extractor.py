"""Extracteur Excel utilisant OpenPyXL."""

import asyncio
from typing import Any

from openpyxl import load_workbook

from app.core.exceptions import ExtractionFailedError
from app.core.logging import get_logger
from app.domain.value_objects.content_metadata import ContentMetadata
from app.domain.value_objects.extraction_result import ExtractionResult, TableBlock
from app.infrastructure.extractors.base import BaseExtractor


class ExcelExtractor(BaseExtractor):
    """Extracteur Excel avec OpenPyXL."""

    def __init__(self, logger: Any = None) -> None:
        """Initialiser l'extracteur."""
        super().__init__(logger or get_logger(__name__))

    async def extract(self, file_path: str) -> ExtractionResult:
        """Extraire le contenu d'un fichier Excel."""
        self._validate_file(file_path)

        try:
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(None, self._extract_sync, file_path)
            return result
        except Exception as e:
            self.logger.exception(f"Erreur lors de l'extraction Excel: {e}")
            raise ExtractionFailedError(f"Échec de l'extraction Excel: {str(e)}")

    def _extract_sync(self, file_path: str) -> ExtractionResult:
        """Extraction synchrone."""
        tables: list[TableBlock] = []

        workbook = load_workbook(file_path, data_only=True, read_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Extraire toutes les données de la feuille
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Filtrer les lignes vides
                    if any(cell is not None and str(cell).strip() for cell in row):
                        rows.append(list(row))

                if rows:
                    # Première ligne comme en-têtes - nettoyer les valeurs None
                    headers = [
                        str(cell).strip() if cell is not None else ""
                        for cell in rows[0]
                    ]
                    # Nettoyer les lignes de données - convertir None en chaîne vide
                    data_rows = [
                        [
                            str(cell).strip() if cell is not None else ""
                            for cell in row
                        ]
                        for row in (rows[1:] if len(rows) > 1 else [])
                    ]

                    metadata = ContentMetadata(
                        order=len(tables),
                        extraction_method="openpyxl",
                        additional_metadata={"sheet_name": sheet_name},
                    )
                    tables.append(TableBlock(headers=headers, rows=data_rows, metadata=metadata))

        finally:
            workbook.close()

        return ExtractionResult(tables=tables)

    async def extract_tables(self, file_path: str) -> list[Any]:
        """Extraire les tableaux (toutes les feuilles)."""
        result = await self.extract(file_path)
        return result.tables

    async def extract_images(self, file_path: str) -> list[Any]:
        """OpenPyXL ne supporte pas l'extraction d'images facilement."""
        return []

    def supports(self, file_type: str) -> bool:
        """Vérifier si le type Excel est supporté."""
        return file_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        ]

